# reports/excel_builder.py
# HBCE — Hybrid Controls Editor
# Professional Excel Report Builder — V0.1.9a-alpha
#
# Uses openpyxl (already in requirements.txt).
# Produces .xlsx files with:
#   Sheet 1 — "Data"    : Formatted table, auto-width columns, frozen header,
#                         alternating stripes, HBCE branding in header rows.
#   Sheet 2 — "Chart"   : Auto-generated chart appropriate to report type:
#                         • alarm_history  → BarChart (priority distribution)
#                         • trend_data     → LineChart (values over time)
#                         • backup_log     → BarChart (status distribution)
#                         • schedule_summary → BarChart (blocks per day)
#                         • point_snapshot → BarChart (objects per type)
#   Sheet 3 — "Summary" : Key statistics table (same as PDF summary block)
#
# Usage:
#   builder = HBCEExcelBuilder("alarm_history", headers, rows, "/path/out.xlsx")
#   builder.build(device_name="FEC2611-0", params={"date_from": "…"})

from __future__ import annotations

from datetime import datetime
from typing import List, Optional

from core.logger import get_logger

logger = get_logger(__name__)

# ── Colour palette ────────────────────────────────────────────────────────────
_ACCENT    = "2B6CB0"   # header fill
_HEADER_FG = "FFFFFF"   # header text
_STRIPE    = "EBF0F7"   # even-row fill
_TITLE_FG  = "1A202C"   # bold title
_META_FG   = "606070"   # grey meta text
_BORDER    = "B8C8DC"   # thin border

# Alarm priority colours (BACnet 1–8)
_PRI_FILLS = {
    "1": "7B0000", "2": "B71C1C", "3": "E53935",
    "4": "FB8C00", "5": "F9A825", "6": "558B2F",
    "7": "1565C0", "8": "37474F",
}

# Which column holds priority values (0-based) — for alarm colour coding
_PRIORITY_COL = {"alarm_history": 5}


class HBCEExcelBuilder:
    """
    Builds a multi-sheet .xlsx report.

    Parameters
    ----------
    report_type : str
    headers : list[str]
    rows : list[list]
    output_path : str
    """

    APP_NAME = "HBCE — Hybrid Controls Editor"

    def __init__(self, report_type: str, headers: List[str],
                 rows: List[list], output_path: str):
        self.report_type = report_type
        self.headers     = headers
        self.rows        = rows
        self.output_path = output_path

    # ── Public API ────────────────────────────────────────────────────────────

    def build(self, report_title: str = "",
              device_name: str = "",
              params: Optional[dict] = None) -> str:
        """
        Generate the Excel file and return the output path.
        Raises RuntimeError if openpyxl is not installed.
        """
        try:
            import openpyxl
            from openpyxl.styles import (
                Font, PatternFill, Alignment, Border, Side,
            )
        except ImportError as e:
            raise RuntimeError(
                "openpyxl is not installed. Run: pip install openpyxl"
            ) from e

        wb = openpyxl.Workbook()
        # Rename default sheet
        ws_data = wb.active
        ws_data.title = "Data"

        title = report_title or self.report_type.replace("_", " ").title()

        self._write_data_sheet(
            ws_data, title, device_name, params or {})

        ws_chart = wb.create_sheet("Chart")
        self._write_chart_sheet(ws_chart, ws_data, title)

        ws_summary = wb.create_sheet("Summary")
        self._write_summary_sheet(ws_summary, title)

        wb.save(self.output_path)
        logger.info(f"Excel report written: {self.output_path}  "
                    f"({len(self.rows)} rows)")
        return self.output_path

    # ── Sheet 1: Data ─────────────────────────────────────────────────────────

    def _write_data_sheet(self, ws, title: str,
                          device_name: str, params: dict):
        try:
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.utils import get_column_letter
        except ImportError:
            return

        def _thin_border():
            s = Side(style="thin", color=_BORDER)
            return Border(left=s, right=s, top=s, bottom=s)

        # ── Branding rows ─────────────────────────────────────────────────────
        ws.append([self.APP_NAME])
        ws["A1"].font      = Font(bold=True, size=12, color=_ACCENT)
        ws["A1"].alignment = Alignment(horizontal="left")

        ts = datetime.now().strftime("%Y-%m-%d  %H:%M:%S")
        ws.append([f"Report: {title}   |   {ts}"
                   + (f"   |   Device: {device_name}" if device_name else "")])
        ws["A2"].font      = Font(size=9, italic=True, color=_META_FG)

        if params:
            parts = [f"{k.replace('_',' ').title()}: {v}"
                     for k, v in params.items() if v]
            if parts:
                ws.append(["Parameters: " + "  ·  ".join(parts)])
                ws["A3"].font = Font(size=8, color=_META_FG)
                ws.append([])           # blank separator
            else:
                ws.append([])
        else:
            ws.append([])

        # ── Column headers ────────────────────────────────────────────────────
        header_row = ws.max_row + 1
        ws.append(self.headers)
        hdr_fill = PatternFill("solid", fgColor=_ACCENT)
        hdr_font = Font(bold=True, color=_HEADER_FG, size=9)
        for col_idx in range(1, len(self.headers) + 1):
            cell = ws.cell(row=header_row, column=col_idx)
            cell.fill      = hdr_fill
            cell.font      = hdr_font
            cell.border    = _thin_border()
            cell.alignment = Alignment(horizontal="center",
                                       vertical="center", wrap_text=True)

        ws.row_dimensions[header_row].height = 18

        # ── Data rows ─────────────────────────────────────────────────────────
        priority_col = _PRIORITY_COL.get(self.report_type, -1)
        stripe_fill  = PatternFill("solid", fgColor=_STRIPE)
        norm_font    = Font(size=8)

        for r_idx, row in enumerate(self.rows):
            ws.append([str(v) if v is not None else "" for v in row])
            actual_row = header_row + 1 + r_idx
            fill = stripe_fill if r_idx % 2 == 0 else PatternFill()

            for col_idx in range(1, len(self.headers) + 1):
                cell = ws.cell(row=actual_row, column=col_idx)
                cell.font      = norm_font
                cell.border    = _thin_border()
                cell.alignment = Alignment(vertical="center")

                # Row stripe (default)
                cell.fill = fill

                # Priority column colour override
                if priority_col >= 0 and col_idx - 1 == priority_col:
                    pri = str(row[priority_col]) if len(row) > priority_col else ""
                    hex_fill = _PRI_FILLS.get(pri)
                    if hex_fill:
                        cell.fill  = PatternFill("solid", fgColor=hex_fill)
                        cell.font  = Font(size=8, bold=True, color="FFFFFF")

        # ── Auto-column width ─────────────────────────────────────────────────
        for col_idx, header in enumerate(self.headers, 1):
            col_letter = get_column_letter(col_idx)
            max_len = len(str(header))
            for r_idx in range(len(self.rows)):
                actual_row = header_row + 1 + r_idx
                val = ws.cell(row=actual_row, column=col_idx).value or ""
                max_len = max(max_len, len(str(val)))
            ws.column_dimensions[col_letter].width = min(max(max_len + 2, 8), 42)

        # ── Freeze header ─────────────────────────────────────────────────────
        ws.freeze_panes = ws.cell(row=header_row + 1, column=1)

        # ── Auto-filter ───────────────────────────────────────────────────────
        if self.rows:
            last_col = get_column_letter(len(self.headers))
            last_row = header_row + len(self.rows)
            ws.auto_filter.ref = (
                f"A{header_row}:{last_col}{last_row}"
            )

    # ── Sheet 2: Chart ────────────────────────────────────────────────────────

    def _write_chart_sheet(self, ws_chart, ws_data, title: str):
        """Generate a chart appropriate to the report type."""
        try:
            from openpyxl.chart import BarChart, LineChart, Reference
            from openpyxl.styles import Font, Alignment
        except ImportError:
            ws_chart["A1"].value = "Charts require openpyxl >= 3.0"
            return

        ws_chart["A1"].value = f"{title} — Chart"
        ws_chart["A1"].font  = Font(bold=True, size=13, color=_ACCENT)
        ws_chart["A1"].alignment = Alignment(horizontal="center")

        rt = self.report_type

        if rt == "trend_data" and len(self.rows) >= 2:
            self._add_trend_chart(ws_chart, title)
        elif rt == "alarm_history":
            self._add_alarm_priority_chart(ws_chart, ws_data, title)
        elif rt == "backup_log":
            self._add_backup_status_chart(ws_chart, ws_data, title)
        elif rt == "schedule_summary":
            self._add_schedule_chart(ws_chart, ws_data, title)
        elif rt == "point_snapshot":
            self._add_object_type_chart(ws_chart, ws_data, title)
        else:
            ws_chart["A3"].value = "No chart available for this report type."

    def _add_trend_chart(self, ws, title: str):
        """Line chart: value over time for trend data."""
        try:
            from openpyxl.chart import LineChart, Reference
        except ImportError:
            return

        if not self.rows:
            return

        # Write mini data table on chart sheet for reference
        start_row = 3
        ws.cell(row=start_row, column=1, value="Timestamp")
        ws.cell(row=start_row, column=2, value="Value")
        max_points = min(len(self.rows), 100)   # cap at 100 for readability
        for i, row in enumerate(self.rows[:max_points]):
            ws.cell(row=start_row + 1 + i, column=1, value=row[0] if row else "")
            try:
                ws.cell(row=start_row + 1 + i, column=2, value=float(row[2]) if len(row) > 2 else 0)
            except (ValueError, TypeError):
                ws.cell(row=start_row + 1 + i, column=2, value=0)

        data_end = start_row + max_points

        chart = LineChart()
        chart.title  = title
        chart.style  = 10
        chart.y_axis.title = "Value"
        chart.x_axis.title = "Sample"
        chart.width  = 20
        chart.height = 12

        values = Reference(ws, min_col=2, min_row=start_row,
                           max_row=data_end)
        chart.add_data(values, titles_from_data=True)
        chart.series[0].graphicalProperties.line.solidFill = _ACCENT
        chart.series[0].graphicalProperties.line.width     = 20000  # 2pt in EMU

        ws.add_chart(chart, "D3")

    def _add_alarm_priority_chart(self, ws, ws_data, title: str):
        """Bar chart: alarm count per priority level."""
        try:
            from openpyxl.chart import BarChart, Reference
            from openpyxl.chart.series import DataPoint
        except ImportError:
            return

        # Tally priority distribution
        pri_counts: dict = {}
        priority_col = _PRIORITY_COL.get(self.report_type, -1)
        if priority_col < 0:
            return

        for row in self.rows:
            if len(row) > priority_col:
                pri = str(row[priority_col])
                pri_counts[pri] = pri_counts.get(pri, 0) + 1

        if not pri_counts:
            return

        labels = sorted(pri_counts.keys(), key=lambda x: int(x) if x.isdigit() else 99)
        start_row = 3
        ws.cell(row=start_row,     column=1, value="Priority")
        ws.cell(row=start_row,     column=2, value="Count")
        ws.cell(row=start_row - 1, column=1, value="Alarm Priority Distribution")

        for i, lbl in enumerate(labels):
            ws.cell(row=start_row + 1 + i, column=1, value=f"P{lbl}")
            ws.cell(row=start_row + 1 + i, column=2, value=pri_counts[lbl])

        data_end = start_row + len(labels)

        chart = BarChart()
        chart.type   = "col"
        chart.style  = 10
        chart.title  = f"{title} — Priority Distribution"
        chart.y_axis.title = "Count"
        chart.x_axis.title = "Priority"
        chart.width  = 16
        chart.height = 10

        data   = Reference(ws, min_col=2, min_row=start_row, max_row=data_end)
        cats   = Reference(ws, min_col=1, min_row=start_row + 1, max_row=data_end)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        chart.series[0].graphicalProperties.solidFill = _ACCENT

        ws.add_chart(chart, "D3")

    def _add_backup_status_chart(self, ws, ws_data, title: str):
        """Bar chart: backup count per status."""
        try:
            from openpyxl.chart import BarChart, Reference
        except ImportError:
            return

        STATUS_COL = 6   # 0-based index in backup_log rows
        counts: dict = {}
        for row in self.rows:
            if len(row) > STATUS_COL:
                s = str(row[STATUS_COL])
                counts[s] = counts.get(s, 0) + 1

        if not counts:
            return

        start_row = 3
        ws.cell(row=start_row - 1, column=1, value="Backup Status Distribution")
        ws.cell(row=start_row,     column=1, value="Status")
        ws.cell(row=start_row,     column=2, value="Count")
        for i, (lbl, cnt) in enumerate(sorted(counts.items())):
            ws.cell(row=start_row + 1 + i, column=1, value=lbl.title())
            ws.cell(row=start_row + 1 + i, column=2, value=cnt)

        data_end = start_row + len(counts)
        chart = BarChart()
        chart.type   = "col"
        chart.style  = 10
        chart.title  = f"{title} — Status Summary"
        chart.width  = 14
        chart.height = 9

        data = Reference(ws, min_col=2, min_row=start_row, max_row=data_end)
        cats = Reference(ws, min_col=1, min_row=start_row + 1, max_row=data_end)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        ws.add_chart(chart, "D3")

    def _add_schedule_chart(self, ws, ws_data, title: str):
        """Bar chart: scheduled hours per day of week."""
        try:
            from openpyxl.chart import BarChart, Reference
        except ImportError:
            return

        DAY_COL  = 2   # 0-based: Mon/Tue/…
        MINS_COL = 5   # duration in minutes
        DAYS_ORDER = ["Mon", "Tue", "Wed", "Thu", "Fri", "Sat", "Sun"]
        mins_per_day: dict = {d: 0 for d in DAYS_ORDER}

        for row in self.rows:
            if len(row) > MINS_COL:
                day = str(row[DAY_COL])
                try:
                    mins = int(row[MINS_COL])
                except (ValueError, TypeError):
                    mins = 0
                if day in mins_per_day:
                    mins_per_day[day] += mins

        start_row = 3
        ws.cell(row=start_row - 1, column=1, value="Scheduled Hours per Day")
        ws.cell(row=start_row,     column=1, value="Day")
        ws.cell(row=start_row,     column=2, value="Scheduled Hours")
        for i, day in enumerate(DAYS_ORDER):
            ws.cell(row=start_row + 1 + i, column=1, value=day)
            ws.cell(row=start_row + 1 + i, column=2,
                    value=round(mins_per_day[day] / 60, 1))

        data_end = start_row + 7
        chart = BarChart()
        chart.type   = "col"
        chart.style  = 10
        chart.title  = "Scheduled Hours per Day of Week"
        chart.y_axis.title = "Hours"
        chart.width  = 16
        chart.height = 10

        data = Reference(ws, min_col=2, min_row=start_row, max_row=data_end)
        cats = Reference(ws, min_col=1, min_row=start_row + 1, max_row=data_end)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        ws.add_chart(chart, "D3")

    def _add_object_type_chart(self, ws, ws_data, title: str):
        """Bar chart: count of each object type in point snapshot."""
        try:
            from openpyxl.chart import BarChart, Reference
        except ImportError:
            return

        TYPE_COL = 1   # 0-based index in point_snapshot rows
        counts: dict = {}
        for row in self.rows:
            if len(row) > TYPE_COL:
                t = str(row[TYPE_COL])
                counts[t] = counts.get(t, 0) + 1

        if not counts:
            return

        start_row = 3
        ws.cell(row=start_row - 1, column=1, value="Points by Object Type")
        ws.cell(row=start_row,     column=1, value="Object Type")
        ws.cell(row=start_row,     column=2, value="Count")
        for i, (lbl, cnt) in enumerate(sorted(counts.items())):
            ws.cell(row=start_row + 1 + i, column=1, value=lbl)
            ws.cell(row=start_row + 1 + i, column=2, value=cnt)

        data_end = start_row + len(counts)
        chart = BarChart()
        chart.type   = "col"
        chart.style  = 10
        chart.title  = "Object Type Distribution"
        chart.width  = 16
        chart.height = 10

        data = Reference(ws, min_col=2, min_row=start_row, max_row=data_end)
        cats = Reference(ws, min_col=1, min_row=start_row + 1, max_row=data_end)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        ws.add_chart(chart, "D3")

    # ── Sheet 3: Summary ──────────────────────────────────────────────────────

    def _write_summary_sheet(self, ws, title: str):
        try:
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        except ImportError:
            return

        def _hdr_cell(row, col, value):
            c = ws.cell(row=row, column=col, value=value)
            c.font      = Font(bold=True, size=10, color="FFFFFF")
            c.fill      = PatternFill("solid", fgColor=_ACCENT)
            c.alignment = Alignment(horizontal="left", vertical="center")

        def _val_cell(row, col, value):
            c = ws.cell(row=row, column=col, value=value)
            c.font      = Font(size=9)
            c.alignment = Alignment(horizontal="left", vertical="center")

        ws.column_dimensions["A"].width = 26
        ws.column_dimensions["B"].width = 40

        ws["A1"].value = f"{self.APP_NAME} — {title}"
        ws["A1"].font  = Font(bold=True, size=12, color=_ACCENT)
        ws["A2"].value = f"Generated: {datetime.now().strftime('%Y-%m-%d  %H:%M:%S')}"
        ws["A2"].font  = Font(size=9, italic=True, color=_META_FG)

        stat_row = 4
        _hdr_cell(stat_row, 1, "Statistic")
        _hdr_cell(stat_row, 2, "Value")

        stats = self._compute_stats()
        for i, (key, val) in enumerate(stats.items()):
            _val_cell(stat_row + 1 + i, 1, key)
            _val_cell(stat_row + 1 + i, 2, str(val))

    def _compute_stats(self) -> dict:
        n  = len(self.rows)
        rt = self.report_type
        out: dict = {"Total rows": n, "Report type": rt.replace("_", " ").title()}

        if rt == "alarm_history":
            pc = _PRIORITY_COL.get(rt, -1)
            active  = sum(1 for r in self.rows if len(r) > 6 and "active" in str(r[6]).lower())
            acked   = sum(1 for r in self.rows if len(r) > 6 and "acked"  in str(r[6]).lower())
            out["Active alarms"]       = active
            out["Acknowledged alarms"] = acked
            out["Cleared alarms"]      = n - active - acked
            if pc >= 0:
                for p in ["1", "2", "3"]:
                    cnt = sum(1 for r in self.rows if len(r) > pc and str(r[pc]) == p)
                    out[f"Priority {p} alarms"] = cnt

        elif rt == "trend_data":
            try:
                vals = [float(r[2]) for r in self.rows if len(r) > 2 and r[2]]
                if vals:
                    out["Min value"] = f"{min(vals):.4f}"
                    out["Max value"] = f"{max(vals):.4f}"
                    out["Avg value"] = f"{sum(vals)/len(vals):.4f}"
            except (ValueError, TypeError):
                pass

        elif rt == "backup_log":
            complete = sum(1 for r in self.rows if len(r) > 6 and r[6] == "complete")
            out["Complete"] = complete
            out["Failed"]   = n - complete

        elif rt == "point_snapshot":
            types: dict = {}
            for r in self.rows:
                t = str(r[1]) if len(r) > 1 else "unknown"
                types[t] = types.get(t, 0) + 1
            for t, cnt in sorted(types.items()):
                out[f"  {t}"] = cnt

        elif rt == "schedule_summary":
            try:
                total_mins = sum(int(r[5]) for r in self.rows if len(r) > 5)
                out["Total scheduled hours"] = f"{total_mins/60:.1f}"
            except (ValueError, TypeError):
                pass

        return out
